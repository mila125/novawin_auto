#Graficos a partir de los dataframes

import openpyxl
import os
import csv
import traceback
from datetime import datetime
from pywinauto import Application, findwindows
import time
import threading
from novawinmng import manejar_novawin, leer_csv_y_crear_dataframe,agregar_csv_a_plantilla_excel, guardar_dataframe_en_ini,generar_nombre_unico,agregar_dataframe_a_excel_sin_borrar,agregar_dataframe_a_nueva_hoja
from pywinauto.keyboard import send_keys
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import subprocess
import openpyxl
import os
import csv
import traceback
from openpyxl import load_workbook
from pandas import ExcelWriter
from openpyxl.drawing.image import Image

def draw_comparison_bar_chart(bjhd, bjha,archivo_planilla):
    # Filtrar las columnas necesarias para ambos dataframes
    radius_bjhd = bjhd['Radius']
    dV_logr_bjhd = bjhd['dV(logr)']
    
    radius_bjha = bjha['Radius']
    dV_logr_bjha = bjha['dV(logr)']
    
    # Asegurarse de que ambas series tengan el mismo tamaño
    min_length = min(len(radius_bjhd), len(radius_bjha))  # Tomar la longitud mínima
    radius_bjhd = radius_bjhd[:min_length]
    dV_logr_bjhd = dV_logr_bjhd[:min_length]
    radius_bjha = radius_bjha[:min_length]
    dV_logr_bjha = dV_logr_bjha[:min_length]

    # Crear el gráfico de barras
    plt.figure(figsize=(12, 6))
    
    # Dibujar las barras para bjhd y bjha
    bar_width = 0.5  # Aumentar el ancho de las barras
    index = range(len(radius_bjhd))
    
    plt.bar(
        index, dV_logr_bjhd, bar_width, 
        label='Desorbcion', edgecolor='blue', fill=False, linewidth=1.5
    )
    plt.bar(
        [i + bar_width for i in index], dV_logr_bjha, bar_width, 
        label='Absorpcion', edgecolor='green', fill=False, linewidth=1.5
    )

    # Etiquetas y título
    plt.xlabel('Radius (Å)')
    plt.ylabel('dV(logr) (cc/Å/g)')
    plt.title('BJH Absorpcion y Desorbcion ')
    
    # Configuración de la cuadrícula
    plt.xticks([i + bar_width / 2 for i in index], radius_bjhd, rotation=90)
    plt.legend()
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    
    # Mostrar el gráfico
    plt.tight_layout()
    plt.show()
    
        # Guardar el gráfico como imagen
    temp_image_path = "grafico_dft.png"
    plt.tight_layout()
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    # Abrir el archivo Excel y cargar la hoja "BJH"
    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "BJH" not in workbook.sheetnames:
            workbook.create_sheet("BJH")
        hoja_hk = workbook["BJH"]

        # Insertar la imagen en la hoja "BJH"
        img = Image(temp_image_path)
        img.anchor = 'G1'  # Posición inicial para insertar la imagen
        hoja_hk.add_image(img)

        # Guardar los cambios en el archivo Excel
        workbook.save(archivo_planilla)
        workbook.close()
        print(f"Gráfico BJH insertado en la hoja 'BJH' del archivo '{archivo_planilla}'.")

        # Eliminar la imagen temporal
        os.remove(temp_image_path)
def draw_DFT(df, archivo_planilla):
    # Filtrar y convertir las columnas necesarias
    half_pore_width = pd.to_numeric(df['Half pore width'], errors='coerce')  # Convertir a numérico
    dVr = df['dV(r)']
    
    # Eliminar filas con valores no convertibles
    valid_indices = ~half_pore_width.isna()
    dVr = dVr[valid_indices]
    half_pore_width = half_pore_width[valid_indices]

    # Crear el gráfico de barras
    plt.figure(figsize=(12, 6))
    bar_width = 0.8  # Aumentar el ancho de las barras

    plt.bar(half_pore_width, dVr, width=bar_width, color='green', alpha=0.7, label='dV(r)')
    
    # Etiquetas y título
    plt.xlabel('Half pore width (Å)')
    plt.ylabel('dV(r) (cc/Å/g)')
    plt.title('DFT Analysis: Half Pore Width vs dV(r)')
    plt.legend()
    
    # Configuración de la cuadrícula
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    
    # Mostrar el gráfico
    plt.tight_layout()
    plt.show()

    # Guardar el gráfico como imagen
    temp_image_path = "grafico_dft.png"
    plt.tight_layout()
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    # Abrir el archivo Excel y cargar la hoja "DFT"
    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "DFT" not in workbook.sheetnames:
            workbook.create_sheet("DFT")
        hoja_hk = workbook["DFT"]

        # Insertar la imagen en la hoja "DFT"
        img = Image(temp_image_path)
        img.anchor = 'G1'  # Posición inicial para insertar la imagen
        hoja_hk.add_image(img)

        # Guardar los cambios en el archivo Excel
        workbook.save(archivo_planilla)
        workbook.close()
        print(f"Gráfico DFT insertado en la hoja 'DFT' del archivo '{archivo_planilla}'.")

        # Eliminar la imagen temporal
        os.remove(temp_image_path)
def draw_HK(df, archivo_planilla):
    # Filtrar y convertir las columnas necesarias
    dVr = df['dV()']
    half_pore_width = pd.to_numeric(df['Half pore width'], errors='coerce')  # Convertir a numérico

    # Eliminar filas con valores no convertibles
    valid_indices = ~half_pore_width.isna()
    dVr = dVr[valid_indices]
    half_pore_width = half_pore_width[valid_indices]

    # Crear el gráfico de barras
    plt.figure(figsize=(14, 6))
    bar_width = 0.8  # Ancho de las barras
    plt.bar(half_pore_width, dVr, width=bar_width, color='green', alpha=0.7, label='dV(r)')

    # Etiquetas y título
    plt.xlabel('Half pore width , A')
    plt.ylabel('dV cc A g')
    plt.title('Gráfico HK')
    plt.xticks(half_pore_width, rotation=90)  # Etiquetas en el eje x
    plt.legend()

    # Guardar el gráfico como imagen
    temp_image_path = "grafico_hk.png"
    plt.tight_layout()
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    # Abrir el archivo Excel y cargar la hoja "HK"
    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "HK" not in workbook.sheetnames:
            workbook.create_sheet("HK")
        hoja_hk = workbook["HK"]

        # Insertar la imagen en la hoja "HK"
        img = Image(temp_image_path)
        img.anchor = 'A1'  # Posición inicial para insertar la imagen
        hoja_hk.add_image(img)

        # Guardar los cambios en el archivo Excel
        workbook.save(archivo_planilla)
        workbook.close()
        print(f"Gráfico HK insertado en la hoja 'HK' del archivo '{archivo_planilla}'.")

        # Eliminar la imagen temporal
        os.remove(temp_image_path)
    else:
        print(f"El archivo '{archivo_planilla}' no existe. No se puede insertar el gráfico.")
def graphs_main(archivo_planilla):
    print("Inicio de graphs_main")
    print(archivo_planilla)
       
    # Crear un diccionario para almacenar los valores extraídos
    valores_nkaffha = {}
    
    # Leer los datos de la hoja 'NKA'
    df_nka = pd.read_excel(archivo_planilla, sheet_name='NKA')
    valores_nka = df_nka.tail(2)  # Obtener los dos últimos valores
    valores_nkaffha['NKA'] = valores_nka
    
    # Leer los datos de la hoja 'FFHA'
    df_ffha = pd.read_excel(archivo_planilla, sheet_name='FFHA')
    valores_ffha = df_ffha.tail(2)  # Obtener los dos últimos valores
    valores_nkaffha['FFHA'] = valores_ffha
    
    # Crear un nuevo DataFrame combinando los valores extraídos
    nuevo_df = pd.concat([
        valores_nka.reset_index(drop=True),
        valores_ffha.reset_index(drop=True)
    ], keys=['NKA', 'FFHA'], names=['Sheet', 'Index'])
    
     # Cargar el archivo Excel con openpyxl
    try:
       book = openpyxl.load_workbook(archivo_planilla)
       # Si la hoja existe, elimínala
       if 'NKAFFHA_valores' in book.sheetnames:
         del book['NKAFFHA_valores']
       book.save(archivo_planilla)
    except FileNotFoundError:
       # Si el archivo no existe, ignora el error
       pass

    # Ahora escribe el DataFrame en el archivo
    with pd.ExcelWriter(archivo_planilla, engine='openpyxl', mode='a') as writer:
      nuevo_df.to_excel(writer, sheet_name='NKAFFHA_valores', index=False)
  
 
    print("Se ha creado la hoja 'NKAFFHA_valores' con los datos de las hojas 'NKA' y 'FFHA'.")
    
    # Leer el archivo Excel
    df = pd.read_excel(archivo_planilla, sheet_name='HK')
    draw_HK(df,archivo_planilla)
    
    # Verificar las primeras filas del archivo para asegurarse de que las columnas existen
    print(df.head())
        # Leer el archivo Excel
    df = pd.read_excel(archivo_planilla, sheet_name='DFT')
    draw_DFT(df,archivo_planilla)
    
    # Verificar las primeras filas del archivo para asegurarse de que las columnas existen
    print(df.head())

    # Leer el archivo Excel
    df_bjha = pd.read_excel(archivo_planilla, sheet_name='BJHA')
    df_bjhd = pd.read_excel(archivo_planilla, sheet_name='BJHD')
   
    draw_comparison_bar_chart(df_bjhd, df_bjha,archivo_planilla)
    # Verificar las primeras filas del archivo para asegurarse de que las columnas existen
    print(df.head())
    # Ejecutar un módulo específico

    result = subprocess.run(["python", "-m", "novarep_ide"], capture_output=True, text=True)
    print("Salida estándar:", result.stdout)
    print("Errores estándar:", result.stderr)

