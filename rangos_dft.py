#Graficos a partir de los dataframes

import openpyxl
import os
import csv
import traceback
from datetime import datetime
from pywinauto import Application, findwindows
import time
import threading
from novawinmng import manejar_novawin, leer_csv_y_crear_dataframe,agregar_csv_a_plantilla_excel, guardar_dataframe_en_ini,generar_nombre_unico,agregar_dataframe_a_excel_sin_borrar,agregar_dataframe_a_nueva_hoja
from pywinauto.keyboard import send_keys
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from pandas import ExcelWriter
from openpyxl import load_workbook
from pandas import ExcelWriter
import subprocess

def procesar_archivo_dft(archivo_ruta_completa, archivo_planilla):
    """
    Procesa un archivo Excel para clasificar los datos en rangos definidos y calcula
    los volúmenes acumulativos de poros. Escribe los resultados en la hoja 'DFT' del archivo.

    Args:
        archivo_ruta_completa (str): Ruta completa del archivo Excel.
        archivo_planilla (str): Nombre del archivo Excel.

    Returns:
        None
    """
    print("Inicio del procesamiento DFT.")
    print(f"Archivo ruta completa: {archivo_ruta_completa}")
    
    # Carga del archivo
    try:
        df = pd.read_excel(archivo_planilla, sheet_name="DFT")
        print("Archivo cargado exitosamente.")
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return

    # Definir rangos
    rangos = {
        "MICRO": (0, 10.5),
        "UNCLASS": (10.49, 18.9),
        "BOTLE": (18.89, 24.8),
        "PLATE": (24.79, 109.7),
        "CYL": (109.69, float('inf'))
    }

    # Inicializar estructuras
    dataframes = {}
    cumulative_pore_volumes = {}

    # Procesar cada rango
    for nombre, (min_val, max_val) in rangos.items():
        try:
            # Convertir a numérico y filtrar NaN
            df['Half pore width'] = pd.to_numeric(df['Half pore width'], errors='coerce')
            df = df.dropna(subset=['Half pore width'])

            # Filtrar por rango
            rango_df = df[(df['Half pore width'] > min_val) & (df['Half pore width'] <= max_val)]
            dataframes[nombre] = rango_df

            # Obtener el valor acumulativo de la última fila
            if not rango_df.empty:
                cumulative_pore_volumes[f"RANGO_CV_{nombre}"] = rango_df['Cumulative Pore Volume'].iloc[-1]
            else:
                cumulative_pore_volumes[f"RANGO_CV_{nombre}"] = None
            
            print(f"Datos filtrados para {nombre}: {len(rango_df)} filas.")
        except Exception as e:
            print(f"Error procesando el rango {nombre}: {e}")

    # Ajustar valores basados en la hoja 'Resultados Tests'
    try:
        libro = openpyxl.load_workbook(archivo_planilla)
        hoja_test = libro['Resultados Tests']
        for row in hoja_test.iter_rows(min_row=2, values_only=True):
            ajustar_valores_por_condiciones(row, cumulative_pore_volumes)
    except Exception as e:
        print(f"Error al procesar la hoja 'Resultados Tests': {e}")

    # Escribir resultados en la hoja 'DFT'
    try:
        hoja_dft = libro['DFT']
        escribir_resultados_en_hoja(hoja_dft, rangos, cumulative_pore_volumes)
        libro.save(archivo_planilla)
        print("Valores escritos en la hoja 'DFT'.")
    except Exception as e:
        print(f"Error al escribir resultados en la hoja 'DFT': {e}")
    
    # Ejecutar módulo externo
    ejecutar_modulo_externo()

def ajustar_valores_por_condiciones(row, cumulative_pore_volumes):
    """
    Ajusta los valores de volúmenes acumulativos en función de las condiciones en una fila.
    """
    if "Hay poros cuello de botella" in row and "Hay poros cilindricos" in row and "Hay poros planos" in row :
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = cumulative_pore_volumes.get("RANGO_CV_BOTLE")
        cumulative_pore_volumes["RANGO_CV_CYL"] = cumulative_pore_volumes.get("RANGO_CV_CYL")
        cumulative_pore_volumes["RANGO_CV_PLATE"] = cumulative_pore_volumes.get("RANGO_CV_PLATE")
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_UNCLASS")
    elif "Hay poros cuello de botella" in row and not "Hay poros cilindricos" in row and "Hay poros planos" in row :
        cumulative_pore_volumes["RANGO_CV_CYL"] = cumulative_pore_volumes.get("RANGO_CV_BOTLE")
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_PLATE"] = cumulative_pore_volumes.get("RANGO_CV_PLATE")+cumulative_pore_volumes.get("RANGO_CV_CYL")
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_UNCLASS")
    elif "Hay poros planos" in row  and not "Hay poros cuello de botella" in row and not "Hay poros cilindricos" in row:
        cumulative_pore_volumes["RANGO_CV_CYL"] = 0 
        cumulative_pore_volumes["RANGO_CV_PLATE"] = cumulative_pore_volumes.get("RANGO_CV_BOTLE") + cumulative_pore_volumes.get("RANGO_CV_CYL") + cumulative_pore_volumes.get("RANGO_CV_PLATE")
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_UNCLASS")
    elif not "Hay poros planos" in row  and not "Hay poros cuello de botella" in row and not "Hay poros cilindricos" in row:
        cumulative_pore_volumes["RANGO_CV_CYL"] = 0 
        cumulative_pore_volumes["RANGO_CV_PLATE"] = 0
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] =  cumulative_pore_volumes.get("RANGO_CV_BOTLE") + cumulative_pore_volumes.get("RANGO_CV_CYL") + cumulative_pore_volumes.get("RANGO_CV_PLATE")
    elif not  "Hay poros cuello de botella" in row and "Hay poros cilindricos" in row and not "Hay poros planos" in row :
        cumulative_pore_volumes["RANGO_CV_CYL"] = cumulative_pore_volumes.get("RANGO_CV_CYL")
        cumulative_pore_volumes["RANGO_CV_PLATE"] =  0
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_BOTLE") + cumulative_pore_volumes.get("RANGO_CV_PLATE")
    elif not  "Hay poros cuello de botella" in row and "Hay poros cilindricos" in row and "Hay poros planos" in row :
        cumulative_pore_volumes["RANGO_CV_CYL"] = 0 
        cumulative_pore_volumes["RANGO_CV_PLATE"] = cumulative_pore_volumes.get("RANGO_CV_BOTLE")  + cumulative_pore_volumes.get("RANGO_CV_PLATE")
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_UNCLASS") 
    elif  "Hay poros cuello de botella" in row and "Hay poros cilindricos" in row and not  "Hay poros planos" in row :
        cumulative_pore_volumes["RANGO_CV_CYL"] = 0 
        cumulative_pore_volumes["RANGO_CV_PLATE"] = 0
        cumulative_pore_volumes["RANGO_CV_BOTLE"] = 0
        cumulative_pore_volumes["RANGO_CV_UNCLASS"] = cumulative_pore_volumes.get("RANGO_CV_PLATE")


def escribir_resultados_en_hoja(hoja_dft, rangos, cumulative_pore_volumes):
    """
    Escribe los resultados de los rangos y volúmenes acumulativos en la hoja 'DFT'.
    """
    # Agregar encabezados si no existen
    if hoja_dft.max_row == 1:
        hoja_dft.append(["Etiqueta", "Rango", "cc/g"])
    
    for nombre, (min_val, max_val) in rangos.items():
        rango_str = f"{min_val} - {max_val}"
        valor = cumulative_pore_volumes.get(f"RANGO_CV_{nombre}", None)
        hoja_dft.append([nombre, rango_str, valor])

def ejecutar_modulo_externo():
    """
    Ejecuta un módulo externo usando subprocess.
    """
    try:
        result = subprocess.run(["python", "-m", "novarep_ide"], capture_output=True, text=True)
        print("Salida estándar:", result.stdout)
        print("Errores estándar:", result.stderr)
    except Exception as e:
        print(f"Error al ejecutar el módulo externo: {e}")

def rangos_dft_main(archivo_ruta_completa, archivo_planilla):
    """
    Procesa un archivo para clasificar los datos en 5 dataframes según rangos definidos
    en la columna 'Half pore width', calcula los valores de 'Cumulative Pore Volume'
    de la última fila de cada rango, y escribe estos valores con sus etiquetas en la hoja 'DFT'.
    
    Args:
    archivo_ruta_completa (str): Ruta completa del archivo que contiene los datos.
    archivo_planilla (str): Tipo de archivo ('csv' o 'excel').
    
    Returns:
    dict: Un diccionario con los 5 dataframes correspondientes a cada rango y los valores
          'Cumulative Pore Volume' para la última fila de cada dataframe.
    """
    print("Inicio de DFT")
    print(f"Archivo ruta completa: {archivo_ruta_completa}")
    
    procesar_archivo_dft(archivo_ruta_completa, archivo_planilla)