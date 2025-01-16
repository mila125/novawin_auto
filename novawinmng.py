from pywinauto import Application
import time
import os
import traceback
import pandas as pd
import configparser
import openpyxl
def agregar_dataframe_a_nueva_hoja(archivo_planilla, nombre_hoja, dataframe):
    """
    Agrega un DataFrame a una nueva hoja en un archivo Excel.
    Si el archivo Excel no existe, se crea.
    Si el nombre de la hoja ya existe, lanza un error.
    
    :param archivo_planilla: Ruta del archivo Excel.
    :param nombre_hoja: Nombre de la nueva hoja.
    :param dataframe: DataFrame que se agregará.
    """
    try:
        # Normalizar ruta según el sistema operativo
        #archivo_planilla = os.path.normpath(archivo_planilla)
        #archivo_planilla = os.path.join(archivo_planilla, "Report.xlsx")
        # Verificar si el archivo Excel existe
        if not os.path.exists(archivo_planilla):
            # Crear un nuevo archivo Excel con la hoja especificada
            with pd.ExcelWriter(archivo_planilla, engine='openpyxl') as writer:
                dataframe.to_excel(writer, sheet_name=nombre_hoja, index=False)
            print(f"Archivo Excel creado con la hoja '{nombre_hoja}': {archivo_planilla}")
        else:
            # Abrir el archivo Excel existente
            with pd.ExcelWriter(archivo_planilla, engine='openpyxl', mode='a') as writer:
                # Verificar si la hoja ya existe
                if nombre_hoja in writer.book.sheetnames:
                    raise ValueError(f"La hoja '{nombre_hoja}' ya existe en el archivo Excel.")
                # Agregar el DataFrame a la nueva hoja
                dataframe.to_excel(writer, sheet_name=nombre_hoja, index=False)
            print(f"Datos agregados exitosamente a la nueva hoja '{nombre_hoja}' en: {archivo_planilla}")
    except Exception as e:
        print(f"Error al agregar el DataFrame a la nueva hoja: {e}")
        raise
def agregar_dataframe_a_excel_sin_borrar(ruta_excel, nuevo_dataframe):
    """
    Agrega un DataFrame a un archivo Excel sin borrar los datos existentes.
    Si el archivo Excel no existe, se crea uno nuevo con los datos del DataFrame.
    """
    try:
        # Invertir las barras en la ruta del archivo
        ruta_excel = ruta_excel.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        ruta_excel = os.path.normpath(ruta_excel)
        ruta_excel = os.path.join(ruta_excel, "Report.xlsx")
        print(ruta_excel)
        # Verificar si el archivo Excel existe
        if not os.path.exists(ruta_excel):
            # Si no existe, crear un archivo Excel nuevo con el DataFrame
            nuevo_dataframe.to_excel(ruta_excel, index=False)
            print(f"Archivo Excel creado: {ruta_excel}")
        else:
            # Si existe, cargar el archivo Excel
            with pd.ExcelFile(ruta_excel) as xl:
                # Leer todas las hojas existentes en el archivo
                hojas = xl.sheet_names
                
                # Si "Reporte" no existe, agregarla
                if "Reporte" not in hojas:
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
                        nuevo_dataframe.to_excel(writer, sheet_name="Reporte", index=False)
                    print(f"Hoja 'Reporte' creada con los nuevos datos en: {ruta_excel}")
                else:
                    # Si "Reporte" ya existe, obtenerla y agregar datos sin borrar
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
                        # Cargar el libro Excel existente
                        wb = openpyxl.load_workbook(ruta_excel)
                        ws = wb["Reporte"]
                        
                        # Encontrar la fila donde agregar nuevos datos (sin sobrescribir)
                        start_row = ws.max_row + 1
                        
                        # Insertar el DataFrame en las filas siguientes
                        for i, row in nuevo_dataframe.iterrows():
                            for j, value in enumerate(row):
                                ws.cell(row=start_row + i, column=j+1, value=value)
                        
                        wb.save(ruta_excel)
                        print(f"Datos agregados exitosamente a la hoja 'Reporte' en: {ruta_excel}")
    except Exception as e:
        print(f"Error al agregar el DataFrame a Excel: {e}")
        raise
import os
from datetime import datetime

def generar_nombre_unico(base_path, namext):
    # Normalizar las barras a formato Unix (/)
    base_path = base_path.replace("\\", "/")
    
    if not base_path.endswith(namext):
        base_path += namext

    # Extraer nombre base y extensión
    name, ext = os.path.splitext(base_path)
    
    # Agregar fecha y hora actual al nombre base
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_with_timestamp = f"{name}_{timestamp}"
    base_path = f"{name_with_timestamp}{ext}"
    
    # Asegurarse de que el nombre sea único
    counter = 1
    while os.path.exists(base_path):
        base_path = f"{name_with_timestamp}_{counter}{ext}"
        counter += 1
    
    # Normalizar las barras de regreso a formato Windows (\)
    return base_path.replace("/", "\\")
    
def manejar_novawin(path_novawin, archivo_qps):
    try:
        # Invertir las barras en la ruta del archivo
        archivo_qps = archivo_qps.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        archivo_qps = os.path.normpath(archivo_qps)

        # Inicializar NovaWin
        app, main_window = inicializar_novawin(path_novawin)

        # Interactuar con NovaWin
        seleccionar_menu(main_window, "File->Open")
        dialog = app.window(class_name="#32770")
        interactuar_con_cuadro_dialogo(dialog, archivo_qps)

        return app, main_window

    except Exception as e:
        print(f"Error al manejar NovaWin: {e}")
        traceback.print_exc()
        raise

def inicializar_novawin(path_novawin):
    try:
        app = Application(backend="uia").start(path_novawin)
        time.sleep(5)  # Esperar que se cargue NovaWin
        main_window = app.window(title_re=".*NovaWin.*")
        return app, main_window
    except Exception as e:
        print(f"Error al inicializar NovaWin: {e}")
        raise

def seleccionar_menu(window, ruta_menu):
    try:
        window.menu_select(ruta_menu)
        time.sleep(2)
    except Exception as e:
        print(f"Error al seleccionar menú '{ruta_menu}': {e}")
        raise

def interactuar_con_cuadro_dialogo(dialog, archivo):
    try:
        edit_box = dialog.child_window(class_name="Edit")
        edit_box.set_edit_text(archivo)
        open_button = dialog.child_window(class_name="Button", found_index=0)
        open_button.click_input()
    except Exception as e:
        print(f"Error al interactuar con el cuadro de diálogo: {e}")
        raise

def leer_csv_y_crear_dataframe(ruta_csv):
    if not os.path.exists(ruta_csv):
        print(f"Archivo CSV no encontrado: {ruta_csv}")
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_csv}")

    try:
        return pd.read_csv(ruta_csv)
    except Exception as e:
        print(f"Error al leer CSV: {e}")
        raise
def agregar_csv_a_plantilla_excel(ruta_csv, ruta_excel,df_csv):
    """
    Agrega el contenido de un CSV a una plantilla Excel (`Reporte.xlsx`).
    Los datos se escriben en las columnas vacías sin borrar el contenido existente.
    """
    try:
        # Invertir las barras en la ruta del archivo
        ruta_excel = ruta_excel.replace("/", "\\")  # Reemplazar barras normales por barras invertidas
        
        # O usar normpath para normalizar la ruta según el sistema operativo
        ruta_excel = os.path.normpath(ruta_excel)
        ruta_excel = os.path.join(ruta_excel, "Report.xlsx")
        print(ruta_excel)
        # Crear archivo Excel si no existe o si el archivo no tiene formato válido
        if not os.path.exists(ruta_excel) or not ruta_excel.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            wb.save(ruta_excel)
            print(f"Archivo Excel creado: {ruta_excel}")

        # Abrir el archivo Excel
        wb = openpyxl.load_workbook(ruta_excel)
        if "Reporte" not in wb.sheetnames:
            ws = wb.create_sheet(title="Reporte")
        else:
            ws = wb["Reporte"]

        # Obtener la fila y columna inicial para insertar los datos
        max_row = ws.max_row
        max_col = ws.max_column

        # Determinar la columna vacía para comenzar a escribir
        start_col = max_col + 1 if max_row > 1 else 1

        # Escribir encabezados si es la primera inserción
        if start_col == 1:
            for col, header in enumerate(df_csv.columns, start=start_col):
                ws.cell(row=1, column=col).value = header

        # Insertar datos en columnas vacías
        for i, row in enumerate(df_csv.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=start_col):
                ws.cell(row=i, column=j).value = value

        # Guardar cambios en el archivo Excel
        wb.save(ruta_excel)
        print(f"Datos del CSV agregados exitosamente a: {ruta_excel}")

    except Exception as e:
        print(f"Error al agregar datos del CSV a la plantilla Excel: {e}")
        raise
def guardar_dataframe_en_ini(df, archivo_ini):
    try:
        config = configparser.ConfigParser()
        for columna in df.columns:
            config[columna] = {f"fila_{i}": str(valor) for i, valor in enumerate(df[columna])}
        with open(archivo_ini, 'w') as archivo:
            config.write(archivo)
        print(f"DataFrame guardado en {archivo_ini}")
    except Exception as e:
        print(f"Error al guardar INI: {e}")
        raise
def close_window_novawin():
    try:
        # Conectar a la ventana de NovaWin
        app = Application(backend='uia').connect(title_re='.*NovaWin.*')
        window = app.window(title_re='.*NovaWin.*')
        # Cerrar la ventana
        window.close()
        print("La ventana de NovaWin ha sido cerrada.")
    except Exception as e:
        print(f"Error al cerrar la ventana de NovaWin: {e}")
def ejecutar_ide():
    subprocess.run(["python", "-m", "novarep_ide"])
